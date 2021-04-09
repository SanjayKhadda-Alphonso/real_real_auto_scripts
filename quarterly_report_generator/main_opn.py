import os
# import xlrd
# import xlwt 
# from xlwt import Workbook 
from openpyxl import load_workbook, Workbook
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def write_2d_list(sheet, row1, col1, twodlist, ind):
	# style = xlwt.XFStyle()

	# # font
	# font = xlwt.Font()
	# font.bold = True
	# style.font = font
	# columnwidth = {}
	# rowx = 1
	# for rowdata in twodlist:
	#     column = col
	#     for colomndata in rowdata:
	#         if column in columnwidth:
	#             columnwidth[column] = max(len(colomndata), len("Conversion Rate"), columnwidth[column])
	#         else:
	#             columnwidth[column] = len(colomndata)
	#         column = column + 1
	#     rowx = rowx + 1	 
	for r in range(len(twodlist)):
		for c in range(len(twodlist[0])):
			# if ind == 0 and r == 0:
			# 	sheet.cell(row = row1 + r, column = col1 + c).value = twodlist[r][c]
			# else:
			cell = sheet.cell(row = row1 + r, column = col1 + c)
			cell.value = twodlist[r][c]["value"]
			cell.number_format = twodlist[r][c]["number_format"]
			cell.font = Font(size=12)
	# column_widths = []
	# for row in twodlist:
	#     for i, cell in enumerate(row):
	#         if len(column_widths) > i:
	#             if len(str(cell["value"])) > column_widths[i]:
	#                 column_widths[i] = len(str(cell["value"]))
	#         else:
	#             column_widths += [len(str(cell["value"]))]

	# for i, column_width in enumerate(column_widths):
	    # sheet.column_dimensions[get_column_letter(i+1)].width = column_width
	# for column, widthvalue in columnwidth.items():
	# 	sheet.col(column).width = (widthvalue + 2) * 256
	# for r in range(1, 10000):
	# 	sheet.row(r).height_mismatch = True
	# 	sheet.row(r).height = 310


def search_string(sheet, string):
	locs = []
	for rowidx in range(1, sheet.max_row+1):
		for colidx in range(1, sheet.max_column+1):
			if sheet.cell(row = rowidx, column = colidx).value == string:
				locs.append((rowidx, colidx))
	return locs

def find_rows_cols(cur_sheet, row, col):
	col_end = col
	row_end = row
	while True:
		try:
			if cur_sheet.cell(row = row, column = col_end).value in [None, ""]:
				break
		except IndexError:
			break					
		col_end = col_end + 1					
	while True:
		try:
			if cur_sheet.cell(row = row_end, column = col).value in [None, ""]:
				break
		except IndexError:
			break
		row_end = row_end + 1

	return row_end, col_end

def main():
	months = {
		"Jan" : "01", 
		"Feb" : "02", 
		"Mar" : "03", 
		"Apr" : "04", 
		"May" : "05", 
		"Jun" : "06", 
		"Jul" : "07", 
		"Aug" : "08", 
		"Sep" : "09", 
		"Oct" : "10", 
		"Nov" : "11", 
		"Dec" : "12", 
	}
	Year = "2020"
	sheet_list = ['TV-Network', 'TV-Daypart', 'TV-Show', 'TV-Creative', 'TV-Content Duration', 'TV-Day Of the Week', 'TV-Creative Daypart', 'TV-Day of the week Daypart', 'TV-Network Daypart', 'TV-NW-DP-Len', 'TV-Network Show', 'TV-Network Day of the Week', 'TV-Recency', 'TV-Frequency', 'OTT-Partner', 'OTT-Partner Provider', 'OTT-Partner Market', 'OTT-Partner Network', 'OTT-Partner Daypart', 'OTT-Partner Day of the Week', 'OTT-Partner Campaign', 'OTT-Daypart', 'OTT-Day Of the Week', 'OTT-Frequency', 'OTT-Recency']


	reports = []
	quarterly_data = []
	data_types = []
	for file in os.listdir("./reports"):
		if file.endswith(".xlsx"):
		    file_name = file.split("-")
		    print(file)
		    start_date = months[file_name[1]] + "/" + file_name[2] + "/" + Year
		    end_date = months[file_name[3]] + "/" + file_name[4] + "/" + Year
		    week = months[file_name[1]] + "/" + file_name[2] + " - " + months[file_name[3]] + "/" + file_name[4]
		    reports.append({"month": file_name[1], "week": week, "start_date": datetime.strptime(start_date, "%m/%d/%Y"), "end_date": datetime.strptime(end_date, "%m/%d/%Y"), "report" : load_workbook("./reports/" + file), "name": file})


	reports.sort(key = lambda x: x["start_date"], reverse = True)
	# print(reports[0]["report"].sheet_names())
	# for report in reports:
	# 	print(report["name"])
	# ws = reports[0]["report"].worksheets[0]
	# print(ws.cell(row = 8, column = 7).value)
	# print(ws.cell(row = 8, column = 7).number_format)
	for sheet in sheet_list:
		sheet_data = {
		"Registration": [], 
		"Purchase": [],
		"Purchase-new": [],
		"Purchase-repeat": []
		}
		first_report = {
		"Registration": True, 
		"Purchase": True,
		"Purchase-new": True,
		"Purchase-repeat": True
		}
		for report in reports:
			# print(report["name"])
			# print("1")
			try:
				cur_sheet = report["report"][sheet]
				evt_coords = search_string(cur_sheet, "Event Name")
				# print(evt_coords)
				for evt in evt_coords:
					# print(evt)
					# print(first_report)
					row, col = evt
					event = cur_sheet.cell(row = row + 1, column = col).value
					# print(event)
					event_data = []
					row_end, col_end = find_rows_cols(cur_sheet, row, col)
					# print("row_end = ", row_end)
					# print("col_end = ", col_end)
				
					if first_report[event] == False:
						row = row + 1
						# print("event1 = ", event)
					# print("row_end = " ,row_end)
					for r in range(row, row_end):
						# print(r)
						row_data = []
						if r == row and first_report[event] == True:
							row_data.append({
								"value" : "Month", 
								"number_format" : 'General'
								})
							row_data.append({
								"value" : "Week of", 
								"number_format" : 'General'
								})
						else:
							row_data.append({
								"value" : report["month"], 
								"number_format" : 'General'
								})
							row_data.append({
								"value" : report["week"], 
								"number_format" : 'General'
								})
						for c in range(col, col_end):
							row_data.append({
								"value" : cur_sheet.cell(r, c).value, 
								"number_format" : cur_sheet.cell(r, c).number_format
								})
							data_types.append(cur_sheet.cell(r, c).number_format)
						event_data.append(row_data)
					
					# print("event = ", event)
					sheet_data[event].append(event_data)
					first_report[event] = False
			except KeyError:
				pass

		quarterly_data.append(sheet_data)		
	
	# print(quarterly_data[3]["Registration"])
	wb = Workbook() 
	
	# borders = xlwt.Borders()
	# borders.bottom = xlwt.Borders.DASHED
	# style.borders = borders
	title = "Alphonso TV Attribution Insights - Hulu Report Jul 13 - Oct 4"
	for i in range(len(sheet_list)):

		worksheet = wb.create_sheet(index = i , title = sheet_list[i])#wb.add_sheet(sheet_list[i])
		start_row = 5
		start_col = 1
		# style = xlwt.XFStyle()

		# # font
		# font = xlwt.Font()
		# font.height = 300
		# font.bold = True
		# style.font = font

		first_cell = worksheet.cell(row = 1, column = 1)
		first_cell.value = title
		first_cell.font = Font(size=20)
		# worksheet.row(0).height_mismatch = True
		# worksheet.row(0).height = 400

		for j in range(len(quarterly_data[i]["Registration"])):
			write_2d_list(worksheet, start_row, start_col, quarterly_data[i]["Registration"][j], j)
			start_row = start_row + len(quarterly_data[i]["Registration"][j])
		start_row = start_row + 3
		# print("sheet = ", sheet_list[i])
		# print(quarterly_data[i]["Registration"][0][0])
		# quarterly_data[i]["Purchase"][0].insert(0, quarterly_data[i]["Registration"][0][0])
		for j in range(len(quarterly_data[i]["Purchase"])):
			write_2d_list(worksheet, start_row, start_col, quarterly_data[i]["Purchase"][j], j)
			start_row = start_row + len(quarterly_data[i]["Purchase"][j])
		start_row = start_row + 3		
		for j in range(len(quarterly_data[i]["Purchase-new"])):
				write_2d_list(worksheet, start_row, start_col, quarterly_data[i]["Purchase-new"][j], j)
				start_row = start_row + len(quarterly_data[i]["Purchase-new"][j])
		start_row = start_row + 3		
		for j in range(len(quarterly_data[i]["Purchase-repeat"])):
				write_2d_list(worksheet, start_row, start_col, quarterly_data[i]["Purchase-repeat"][j], j)
				start_row = start_row + len(quarterly_data[i]["Purchase-repeat"][j])	
		start_row = start_row + 3
		
		ws = worksheet
		dims = {}
		for row in ws.rows:
		    for cell in row:
		        if cell.value:
		        	if cell.number_format == '#,##0':
		        		dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))*2))  
		        	elif cell.number_format == '0.00%':
		        		dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
		        	else:
		        		dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
		for col, value in dims.items():
		    ws.column_dimensions[colnum_string(col)].width = value
		ws.column_dimensions['A'].width = 10
	# print(list(set(data_types)))
	#['0.00%', '#,##0', 'General']
	wb.save("./TRR_hulu_consolidated.xlsx")                            

main()
